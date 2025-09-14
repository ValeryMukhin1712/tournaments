import os
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from flask import send_file, flash
import logging

logger = logging.getLogger(__name__)

def get_downloads_folder():
    """Определяет папку Загрузки/Downloads пользователя"""
    try:
        # Получаем путь к домашней папке пользователя
        home = os.path.expanduser("~")
        
        # Проверяем разные варианты папки загрузок
        possible_downloads = [
            os.path.join(home, "Downloads"),  # Windows (английский)
            os.path.join(home, "Загрузки"),   # Windows (русский)
            os.path.join(home, "downloads"),  # Linux (английский)
            os.path.join(home, "загрузки"),   # Linux (русский)
        ]
        
        # Ищем существующую папку
        for downloads_path in possible_downloads:
            if os.path.exists(downloads_path) and os.path.isdir(downloads_path):
                return downloads_path
        
        # Если не найдена, создаем папку Downloads в домашней директории
        downloads_path = os.path.join(home, "Downloads")
        os.makedirs(downloads_path, exist_ok=True)
        return downloads_path
        
    except Exception as e:
        logger.warning(f"Не удалось определить папку загрузок: {e}")
        # Fallback на текущую директорию
        return os.getcwd()

def create_excel_export(tournament, participants, matches, statistics, positions):
    """Создает Excel файл с данными турнира"""
    try:
        # Создаем рабочую книгу
        wb = Workbook()
        
        # Удаляем стандартный лист
        wb.remove(wb.active)
        
        # Создаем листы
        create_tournament_info_sheet(wb, tournament)
        create_standings_sheet(wb, participants, statistics, positions)
        create_chessboard_sheet(wb, tournament, participants, matches)
        create_schedule_sheet(wb, matches, participants)
        
        # Создаем имя файла
        tournament_name = tournament.name.replace(' ', '_').replace('/', '_')
        date_str = datetime.now().strftime('%Y%m%d_%H%M')
        filename = f"{tournament_name}_{date_str}.xlsx"
        
        # Используем папку загрузок пользователя
        downloads_dir = get_downloads_folder()
        filepath = os.path.join(downloads_dir, filename)
        
        # Сохраняем файл
        wb.save(filepath)
        
        logger.info(f"Excel файл создан: {filepath}")
        return filepath, filename
        
    except Exception as e:
        logger.error(f"Ошибка при создании Excel файла: {e}")
        raise e

def create_tournament_info_sheet(wb, tournament):
    """Создает лист с информацией о турнире"""
    ws = wb.create_sheet("Информация о турнире")
    
    # Заголовок
    ws['A1'] = "ИНФОРМАЦИЯ О ТУРНИРЕ"
    ws['A1'].font = Font(size=16, bold=True)
    ws.merge_cells('A1:D1')
    
    # Информация о турнире
    info_data = [
        ("Название турнира:", tournament.name),
        ("Тип спорта:", getattr(tournament, 'sport_type', 'Не указано')),
        ("Дата начала:", tournament.start_date.strftime('%d.%m.%Y')),
        ("Дата окончания:", tournament.end_date.strftime('%d.%m.%Y')),
        ("Время начала:", tournament.start_time.strftime('%H:%M') if tournament.start_time else "Не указано"),
        ("Время окончания:", tournament.end_time.strftime('%H:%M') if tournament.end_time else "Не указано"),
        ("Максимум участников:", tournament.max_participants),
        ("Количество площадок:", tournament.court_count),
        ("Длительность матча (мин):", tournament.match_duration),
        ("Длительность перерыва (мин):", tournament.break_duration),
        ("Сетов для победы:", tournament.sets_to_win),
        ("Очков для победы в сете:", tournament.points_to_win),
        ("Очки за победу:", tournament.points_win),
        ("Очки за ничью:", tournament.points_draw),
        ("Очки за поражение:", tournament.points_loss),
        ("Статус:", getattr(tournament, 'status', 'Не указано')),
        ("Дата создания:", tournament.created_at.strftime('%d.%m.%Y %H:%M') if tournament.created_at else "Не указано"),
    ]
    
    row = 3
    for label, value in info_data:
        ws[f'A{row}'] = label
        ws[f'B{row}'] = value
        ws[f'A{row}'].font = Font(bold=True)
        row += 1
    
    # Настройка ширины колонок
    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 30

def create_standings_sheet(wb, participants, statistics, positions):
    """Создает лист с турнирной таблицей"""
    ws = wb.create_sheet("Турнирная таблица")
    
    # Заголовки
    headers = ["Место", "Участник", "Игры", "Победы", "Поражения", "Ничьи", "Очки", "Разность мячей"]
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
        cell.alignment = Alignment(horizontal="center")
    
    # Данные участников
    row = 2
    for participant in participants:
        stats = statistics.get(participant.id, {})
        position = positions.get(participant.id, 1)
        
        ws.cell(row=row, column=1, value=position)
        ws.cell(row=row, column=2, value=participant.name)
        ws.cell(row=row, column=3, value=stats.get('games', 0))
        ws.cell(row=row, column=4, value=stats.get('wins', 0))
        ws.cell(row=row, column=5, value=stats.get('losses', 0))
        ws.cell(row=row, column=6, value=stats.get('draws', 0))
        ws.cell(row=row, column=7, value=stats.get('points', 0))
        ws.cell(row=row, column=8, value=stats.get('goal_difference', 0))
        
        # Выделяем лидера
        if position == 1:
            for col in range(1, 9):
                ws.cell(row=row, column=col).fill = PatternFill(start_color="FFD700", end_color="FFD700", fill_type="solid")
        
        row += 1
    
    # Настройка ширины колонок
    for col in range(1, 9):
        ws.column_dimensions[get_column_letter(col)].width = 15

def create_chessboard_sheet(wb, tournament, participants, matches):
    """Создает лист с турнирной таблицей (шахматка)"""
    ws = wb.create_sheet("Турнирная таблица (шахматка)")
    
    # Сортируем участников по имени
    sorted_participants = sorted(participants, key=lambda p: p.name)
    
    # Создаем словарь матчей для быстрого поиска
    matches_dict = {}
    for match in matches:
        key = tuple(sorted([match.participant1_id, match.participant2_id]))
        matches_dict[key] = match
    
    # Заголовки
    ws.cell(row=1, column=1, value="Участник")
    ws.cell(row=1, column=1).font = Font(bold=True)
    
    for col, participant in enumerate(sorted_participants, 2):
        ws.cell(row=1, column=col, value=participant.name[:15])
        ws.cell(row=1, column=col).font = Font(bold=True)
    
    # Заполняем таблицу
    for row, p1 in enumerate(sorted_participants, 2):
        ws.cell(row=row, column=1, value=p1.name)
        ws.cell(row=row, column=1).font = Font(bold=True)
        
        for col, p2 in enumerate(sorted_participants, 2):
            if p1.id == p2.id:
                ws.cell(row=row, column=col, value="—")
            else:
                # Ищем матч между участниками
                key = tuple(sorted([p1.id, p2.id]))
                match = matches_dict.get(key)
                
                if match and match.status == 'завершен':
                    # Определяем счёт с точки зрения p1
                    if match.participant1_id == p1.id:
                        score = f"{getattr(match, 'sets_won_1', 0)}:{getattr(match, 'sets_won_2', 0)}"
                        is_winner = getattr(match, 'sets_won_1', 0) > getattr(match, 'sets_won_2', 0)
                    else:
                        score = f"{getattr(match, 'sets_won_2', 0)}:{getattr(match, 'sets_won_1', 0)}"
                        is_winner = getattr(match, 'sets_won_2', 0) > getattr(match, 'sets_won_1', 0)
                    
                    cell = ws.cell(row=row, column=col, value=score)
                    cell.alignment = Alignment(horizontal="center")
                    
                    # Цветовое выделение
                    if is_winner:
                        cell.fill = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")
                    else:
                        cell.fill = PatternFill(start_color="FFB6C1", end_color="FFB6C1", fill_type="solid")
                elif match:
                    # Матч запланирован или в процессе
                    time_str = ""
                    if match.match_time:
                        time_str = match.match_time.strftime('%H:%M')
                    if match.court_number:
                        time_str += f" Пл.{match.court_number}"
                    ws.cell(row=row, column=col, value=time_str or "Запланирован")
                else:
                    ws.cell(row=row, column=col, value="")
    
    # Настройка ширины колонок
    ws.column_dimensions['A'].width = 20
    for col in range(2, len(sorted_participants) + 2):
        ws.column_dimensions[get_column_letter(col)].width = 12

def create_schedule_sheet(wb, matches, participants):
    """Создает лист с расписанием матчей"""
    ws = wb.create_sheet("Расписание матчей")
    
    # Заголовки
    headers = ["№", "Дата", "Время", "Площадка", "Участник 1", "Участник 2", "Счёт", "Статус"]
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
        cell.alignment = Alignment(horizontal="center")
    
    # Создаем словарь участников для быстрого поиска
    participants_dict = {p.id: p for p in participants}
    
    # Сортируем матчи по дате и времени
    sorted_matches = sorted(matches, key=lambda m: (m.match_date or datetime.min.date(), m.match_time or datetime.min.time()))
    
    # Данные матчей
    row = 2
    for i, match in enumerate(sorted_matches, 1):
        participant1 = participants_dict.get(match.participant1_id)
        participant2 = participants_dict.get(match.participant2_id)
        
        ws.cell(row=row, column=1, value=i)
        ws.cell(row=row, column=2, value=match.match_date.strftime('%d.%m.%Y') if match.match_date else "")
        ws.cell(row=row, column=3, value=match.match_time.strftime('%H:%M') if match.match_time else "")
        ws.cell(row=row, column=4, value=match.court_number or "")
        ws.cell(row=row, column=5, value=participant1.name if participant1 else f"ID:{match.participant1_id}")
        ws.cell(row=row, column=6, value=participant2.name if participant2 else f"ID:{match.participant2_id}")
        
        # Счёт
        if match.status == 'завершен':
            score = f"{getattr(match, 'sets_won_1', 0)}:{getattr(match, 'sets_won_2', 0)}"
            ws.cell(row=row, column=7, value=score)
        else:
            ws.cell(row=row, column=7, value="—")
        
        # Статус
        status_colors = {
            'завершен': '90EE90',
            'играют': 'FFFF99',
            'запланирован': 'E6E6FA'
        }
        cell = ws.cell(row=row, column=8, value=match.status)
        if match.status in status_colors:
            cell.fill = PatternFill(start_color=status_colors[match.status], end_color=status_colors[match.status], fill_type="solid")
        
        row += 1
    
    # Настройка ширины колонок
    column_widths = [5, 12, 10, 10, 20, 20, 10, 15]
    for col, width in enumerate(column_widths, 1):
        ws.column_dimensions[get_column_letter(col)].width = width
